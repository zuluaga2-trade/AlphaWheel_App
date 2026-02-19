# AlphaWheel Pro - Bitácora: PDF, Excel, CSV, Tax Efficiency, comentarios
# Reportes flexibles por rango de fechas; hojas bien elaboradas para combinar con otras
import io
from datetime import datetime
from typing import List, Dict, Any, Optional

import pandas as pd
import streamlit as st
from database import db
from engine.calculations import round2, safe_float
from business.wheel import get_campaign_root_id, get_campaign_start_date


def _trades_to_dataframe(trades: List[Dict], account_name: str) -> pd.DataFrame:
    """Convierte lista de trades a DataFrame con columnas estándar y 2 decimales."""
    if not trades:
        return pd.DataFrame()
    df = pd.DataFrame(trades)
    for col in ("price", "strike"):
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round2(x) if x is not None else None)
    df["account_name"] = account_name
    return df


def _get_trades_for_report(
    account_id: int,
    date_from: str,
    date_to: str,
    ticker: Optional[str] = None,
    strategy: Optional[str] = None,
    status: Optional[str] = None,
) -> List[Dict]:
    """
    Trades que abrieron O cerraron en el rango (para ver campaña completa).
    - Apertura en rango: trade_date >= date_from AND trade_date <= date_to
    - Cierre en rango: closed_date >= date_from AND closed_date <= date_to
    Así si cierras por recompra en febrero, el trade sale en el reporte de febrero.

    Para que el comportamiento sea idéntico en SQLite y PostgreSQL y evitar problemas
    con diferencias en SQL, obtenemos todos los trades de la cuenta y filtramos en Python.
    """
    # Obtener todos los trades de la cuenta
    try:
        all_trades = db.get_trades_by_account(account_id)
    except Exception:
        all_trades = []

    rows: List[Dict] = []
    try:
        d_from = datetime.fromisoformat(date_from).date()
        d_to = datetime.fromisoformat(date_to).date()
    except Exception:
        d_from = None
        d_to = None

    for t in all_trades:
        # Filtros opcionales por ticker / estrategia / estado
        if ticker and (t.get("ticker") or "").strip().upper() != (ticker or "").strip().upper():
            continue
        if strategy and (t.get("strategy_type") or "").strip().upper() != (strategy or "").strip().upper():
            continue
        if status and (t.get("status") or "").strip().upper() != (status or "").strip().upper():
            continue

        td_str = (t.get("trade_date") or "")[:10]
        cd_str = (t.get("closed_date") or "")[:10]
        try:
            td = datetime.fromisoformat(td_str).date() if td_str else None
        except Exception:
            td = None
        try:
            cd = datetime.fromisoformat(cd_str).date() if cd_str else None
        except Exception:
            cd = None

        # Condiciones de rango
        cond_open = (d_from is None or d_to is None) or (td is not None and d_from <= td <= d_to)
        cond_close = (d_from is None or d_to is None) or (cd is not None and d_from <= cd <= d_to)
        if not (cond_open or cond_close):
            continue

        rows.append(dict(t))

    # Enriquecer: campaña + total USD (convención opciones: 1 contrato = 100, total = precio × 100 × contratos)
    for r in rows:
        root_id = get_campaign_root_id(account_id, r["trade_id"])
        r["campaign_root_id"] = root_id
        r["campaign_start_date"] = (
            get_campaign_start_date(account_id, r["trade_id"]) if root_id else None
        )
        atype = (r.get("asset_type") or "").strip().upper()
        qty = int(r.get("quantity") or 0)
        # No usar safe_float(price): redondea a 2 decimales y anula débitos pequeños (ej. 0.02 → price -0.0001 → 0)
        try:
            p_raw = float(r.get("price")) if r.get("price") is not None else 0.0
        except (TypeError, ValueError):
            p_raw = 0.0
        # Total en USD: opciones = precio × 100 × contratos (prima positiva, débito negativo); acciones = precio × cantidad
        mult = 100 if atype == "OPTION" else 1
        r["total_usd"] = round2(p_raw * qty * mult)

        # Recompra: trade de cierre. Débito = precio por acción × 100 × contratos (como la prima). Preferir BD; si no, derivar.
        entry = (r.get("entry_type") or "").strip().upper()
        is_closing = (
            entry == "CLOSING"
            or (r.get("parent_trade_id") and atype == "OPTION" and p_raw <= 0)
            or (atype == "OPTION" and p_raw < 0)
        )
        if is_closing and atype == "OPTION":
            r["close_type"] = r.get("close_type") or "buyback"
            if r.get("buyback_debit") is not None:
                try:
                    r["buyback_debit"] = round2(float(r["buyback_debit"]))
                except (TypeError, ValueError):
                    r["buyback_debit"] = round2(abs(p_raw) * qty * 100) if qty else 0.0
            else:
                r["buyback_debit"] = round2(abs(p_raw) * qty * 100) if qty else round2(abs(r["total_usd"]))
            # Para que el neto del periodo sea correcto: recompra resta (total_usd negativo)
            r["total_usd"] = -round2(float(r["buyback_debit"]))
    return rows


@st.cache_data(ttl=120, show_spinner=False)
def count_trades_for_account(account_id: int) -> int:
    """Número total de trades de la cuenta (cualquier fecha). Para mensajes cuando el rango devuelve vacío."""
    conn = db.get_conn()
    try:
        cur = conn.execute("SELECT COUNT(*) AS n FROM Trade WHERE account_id = ?", (account_id,))
        row = cur.fetchone()
        return int(row.get("n", 0) or 0) if row else 0
    finally:
        conn.close()


@st.cache_data(ttl=120, show_spinner=False)
def get_trade_filter_options(account_id: int) -> Dict[str, List[str]]:
    """
    Opciones ligeras para los filtros de reporte (tickers y estrategias).
    Cache 2 min para no repetir cálculo al cambiar solo filtros.
    """
    try:
        trades = db.get_trades_by_account(account_id) or []
    except Exception:
        trades = []
    tickers = sorted({(t.get("ticker") or "").strip().upper() for t in trades if t.get("ticker")})
    strategies = sorted({(t.get("strategy_type") or "").strip().upper() for t in trades if t.get("strategy_type")})
    return {"tickers": tickers, "strategies": strategies}


def get_trades_for_report(
    account_id: int,
    date_from: str,
    date_to: str,
    ticker: Optional[str] = None,
    strategy: Optional[str] = None,
    status: Optional[str] = None,
) -> List[Dict]:
    """
    API pública: lista de trades en el rango para preview y reportes.
    Sin caché para que la web siempre muestre datos actuales (evitar resultado vacío cacheado).
    """
    return _get_trades_for_report(account_id, date_from, date_to, ticker, strategy, status)


def export_trades_csv(
    account_id: int,
    date_from: str,
    date_to: str,
    account_name: str = "",
) -> str:
    """Exporta trades del rango de fechas a CSV (contenido como string)."""
    rows = _get_trades_for_report(account_id, date_from, date_to)
    df = _trades_to_dataframe(rows, account_name)
    if df.empty:
        return ""
    buf = io.StringIO()
    df.to_csv(buf, index=False, encoding="utf-8")
    return buf.getvalue()


def export_trades_excel(
    account_id: int,
    date_from: str,
    date_to: str,
    account_name: str = "",
):
    """
    Exporta a Excel: hoja Trades (columnas claras para juntar con otras) y hoja Resumen.
    Flexible por rango de fechas; formato listo para operaciones.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    rows = _get_trades_for_report(account_id, date_from, date_to)
    if not rows:
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(["Fecha", "Ticker", "Activo_tipo", "Estrategia", "Cantidad", "Precio_por_accion", "Total_USD", "Strike", "Fecha_exp", "Estado", "Fecha_cierre", "Tipo_entrada", "Debito_recompra", "Comentario"])
        wb.save(buf)
        return buf.getvalue()

    # Hoja Trades: Total_USD = precio × 100 × contratos (opciones) o precio × cantidad (stock). Débito resta (negativo en Total_USD).
    data = []
    for r in rows:
        total_usd = r.get("total_usd")
        if total_usd is None:
            atype = (r.get("asset_type") or "").strip().upper()
            qty = int(r.get("quantity") or 0)
            p = safe_float(r.get("price"))
            total_usd = round2(p * qty * (100 if atype == "OPTION" else 1))
        is_recompra = (r.get("close_type") or "").lower() == "buyback" or ((r.get("entry_type") or "").upper() == "CLOSING" and (r.get("asset_type") or "").upper() == "OPTION")
        debito = r.get("buyback_debit") if is_recompra else ""
        if debito is not None and debito != "":
            debito = round2(float(debito))
        data.append([
            str(r.get("trade_date", ""))[:10],
            str(r.get("ticker", "")),
            str(r.get("asset_type", "")),
            str(r.get("strategy_type", "")),
            int(r.get("quantity", 0)),
            round2(r.get("price")),
            total_usd,
            round2(r.get("strike")) if r.get("strike") is not None else "",
            str(r.get("expiration_date") or "")[:10],
            str(r.get("status", "")),
            str(r.get("closed_date") or "")[:10],
            str(r.get("entry_type", "")),
            debito if debito != "" else "",
            (r.get("comment") or "")[:200],
        ])
    df_trades = pd.DataFrame(data, columns=[
        "Fecha", "Ticker", "Activo_tipo", "Estrategia", "Cantidad", "Precio_por_accion", "Total_USD",
        "Strike", "Fecha_exp", "Estado", "Fecha_cierre", "Tipo_entrada", "Debito_recompra", "Comentario"
    ])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_trades.to_excel(writer, sheet_name="Trades", index=False)
        # Resumen: Prima_total = suma de Total_USD (débitos ya vienen negativos, netean)
        resumen_ticker = df_trades.groupby("Ticker").apply(
            lambda g: pd.Series({
                "Cantidad_trades": len(g),
                "Prima_neto_USD": g["Total_USD"].sum(),
            })
        ).reset_index()
        resumen_estrategia = df_trades.groupby("Estrategia").agg(
            Cantidad_trades=("Estrategia", "count"),
        ).reset_index()
        resumen_ticker.to_excel(writer, sheet_name="Resumen_por_ticker", index=False)
        resumen_estrategia.to_excel(writer, sheet_name="Resumen_por_estrategia", index=False)

    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    thin = Side(style="thin")
    header_font = Font(bold=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(12, min(24, len(str(ws.cell(1, col).value or "")) + 2))
    buf2 = io.BytesIO()
    wb.save(buf2)
    return buf2.getvalue()


def export_trades_pdf(
    account_id: int,
    date_from: str,
    date_to: str,
    account_name: str = "",
) -> bytes:
    """Exporta bitácora a PDF (rango de fechas flexible). Columnas claras para bitácora exacta."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        try:
            from fpdf import FPDF
            return _pdf_fpdf(account_id, date_from, date_to, account_name)
        except ImportError:
            return b""

    rows = _get_trades_for_report(account_id, date_from, date_to)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("AlphaWheel Pro — Bitácora de Trades", styles["Title"]))
    story.append(Paragraph(f"Cuenta: {account_name} | Desde: {date_from} | Hasta: {date_to}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if not rows:
        story.append(Paragraph("No hay trades en este rango de fechas.", styles["Normal"]))
    else:
        # Prima/Débito en total USD: precio × 100 × contratos (misma fórmula; débito resta)
        data = [["Fecha", "Ticker", "Tipo", "Estrategia", "Cant", "Prima (USD)", "Débito (USD)", "Strike", "Exp", "Estado", "Cierre", "Entrada", "Comentario"]]
        for r in rows:
            total_usd = r.get("total_usd")
            if total_usd is None:
                atype = (r.get("asset_type") or "").strip().upper()
                qty = int(r.get("quantity") or 0)
                total_usd = round2(safe_float(r.get("price")) * qty * (100 if atype == "OPTION" else 1))
            is_recompra = (r.get("close_type") or "").lower() == "buyback" or ((r.get("entry_type") or "").upper() == "CLOSING" and (r.get("asset_type") or "").upper() == "OPTION")
            prima_cell = str(round2(total_usd)) if not is_recompra and total_usd is not None else "-"
            debito_cell = str(round2(r.get("buyback_debit"))) if is_recompra and r.get("buyback_debit") is not None else "-"
            data.append([
                str(r.get("trade_date", ""))[:10],
                str(r.get("ticker", "")),
                str(r.get("asset_type", "")),
                str(r.get("strategy_type", "")),
                str(r.get("quantity", "")),
                prima_cell,
                debito_cell,
                str(round2(r.get("strike")) if r.get("strike") is not None else "-"),
                str(r.get("expiration_date") or "-")[:10],
                str(r.get("status", "")),
                str(r.get("closed_date") or "-")[:10],
                str(r.get("entry_type", "")),
                (r.get("comment") or "")[:25],
            ])
        t = Table(data, colWidths=[48, 40, 32, 48, 24, 44, 44, 40, 48, 32, 48, 40, 72])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
    doc.build(story)
    return buf.getvalue()


def _pdf_fpdf(account_id: int, date_from: str, date_to: str, account_name: str) -> bytes:
    from fpdf import FPDF
    rows = _get_trades_for_report(account_id, date_from, date_to)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "AlphaWheel Pro - Bitacora", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Cuenta: {account_name} | {date_from} a {date_to}", ln=True)
    pdf.ln(6)
    if rows:
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(20, 6, "Fecha"); pdf.cell(16, 6, "Ticker"); pdf.cell(12, 6, "Tipo"); pdf.cell(12, 6, "Cant"); pdf.cell(18, 6, "Prima USD"); pdf.cell(18, 6, "Debito USD"); pdf.cell(16, 6, "Strike"); pdf.cell(20, 6, "Exp"); pdf.cell(12, 6, "Estado"); pdf.cell(20, 6, "Cierre"); pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        for r in rows:
            total_usd = r.get("total_usd")
            if total_usd is None:
                atype = (r.get("asset_type") or "").strip().upper()
                qty = int(r.get("quantity") or 0)
                total_usd = round2(safe_float(r.get("price")) * qty * (100 if atype == "OPTION" else 1))
            is_recompra = (r.get("close_type") or "").lower() == "buyback" or ((r.get("entry_type") or "").upper() == "CLOSING" and (r.get("asset_type") or "").upper() == "OPTION")
            prima_str = str(round2(total_usd)) if not is_recompra else "-"
            debito_str = str(round2(r.get("buyback_debit"))) if is_recompra and r.get("buyback_debit") is not None else "-"
            pdf.cell(20, 5, str(r.get("trade_date", ""))[:10]); pdf.cell(16, 5, str(r.get("ticker", ""))); pdf.cell(12, 5, str(r.get("asset_type", ""))); pdf.cell(12, 5, str(r.get("quantity", ""))); pdf.cell(18, 5, prima_str); pdf.cell(18, 5, debito_str); pdf.cell(16, 5, str(round2(r.get("strike")) if r.get("strike") else "-")); pdf.cell(20, 5, str(r.get("expiration_date") or "-")[:10]); pdf.cell(12, 5, str(r.get("status", ""))); pdf.cell(20, 5, str(r.get("closed_date") or "-")[:10]); pdf.ln()
    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


def tax_efficiency_summary(
    account_id: int,
    date_from: str,
    date_to: str,
) -> Dict[str, Any]:
    """
    Resumen Tax Efficiency: ganancias/pérdidas cerradas en el rango.
    Sin caché para que la web muestre datos actuales.
    """
    conn = db.get_conn()
    try:
        # Trades cerrados en el rango. No usar close_type/buyback_debit en SELECT (pueden no existir en BD antigua).
        cur = conn.execute(
            """SELECT trade_id, ticker, asset_type, strategy_type, quantity, price, strike,
                      trade_date, closed_date, entry_type
               FROM Trade
               WHERE account_id = ?
                 AND status = 'CLOSED'
                 AND closed_date >= ?
                 AND closed_date <= ?
               ORDER BY closed_date""",
            (account_id, date_from, date_to),
        )
        rows = [dict(r) for r in cur.fetchall()]

        # Capital de referencia de la cuenta (para ratios)
        cur_acc = conn.execute(
            "SELECT cap_total FROM Account WHERE account_id = ?",
            (account_id,),
        )
        acc_row = cur_acc.fetchone()
        cap_total = float(acc_row["cap_total"]) if acc_row and acc_row["cap_total"] is not None else 0.0
    finally:
        conn.close()

    total_realized = 0.0
    by_ticker: Dict[str, float] = {}
    by_strategy: Dict[str, float] = {}
    closed_by_strategy: Dict[str, int] = {}

    for r in rows:
        # P&L por trade: prima recibida (positivo) o débito recompra (negativo). Preferir buyback_debit si está en BD (precisión).
        mult = 100 if r["asset_type"] == "OPTION" else 1
        if (r.get("close_type") or "").lower() == "buyback" and r.get("buyback_debit") is not None:
            amt = -float(r["buyback_debit"])
        else:
            amt = float(r["price"]) * int(r["quantity"]) * mult
        total_realized += amt

        tk = r["ticker"]
        strat = r.get("strategy_type") or "OTHER"
        by_ticker[tk] = by_ticker.get(tk, 0.0) + amt
        by_strategy[strat] = by_strategy.get(strat, 0.0) + amt
        closed_by_strategy[strat] = closed_by_strategy.get(strat, 0) + 1

    # Restar comisiones y fees por campaña (una vez por campaña cerrada en el rango)
    campaign_roots = set()
    for r in rows:
        root_id = get_campaign_root_id(account_id, r["trade_id"])
        if root_id:
            campaign_roots.add(root_id)
    for root_id in campaign_roots:
        adj = db.get_campaign_adjustment(account_id, root_id)
        total_realized -= safe_float(adj.get("commissions", 0)) + safe_float(adj.get("fees", 0))

    total_realized = round2(total_realized)
    realized_pct_of_capital = round2((total_realized / cap_total * 100) if cap_total else 0.0)

    # Anualización aproximada sobre el rango de fechas del reporte
    try:
        d0 = datetime.fromisoformat(date_from)
        d1 = datetime.fromisoformat(date_to)
        days = max(1, (d1.date() - d0.date()).days or 1)
    except Exception:
        days = 1
    realized_ann_pct = round2((realized_pct_of_capital / days) * 365.0) if days else 0.0

    return {
        "total_realized_gain_loss": total_realized,
        "closed_trades_count": len(rows),
        "by_ticker": {k: round2(v) for k, v in by_ticker.items()},
        "by_strategy": {k: round2(v) for k, v in by_strategy.items()},
        "closed_trades_by_strategy": closed_by_strategy,
        "capital_reference": round2(cap_total),
        "realized_pct_of_capital": realized_pct_of_capital,
        "realized_ann_pct": realized_ann_pct,
        "date_from": date_from,
        "date_to": date_to,
    }
